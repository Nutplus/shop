import openpyxl

path = 'excel_data/user_info.xlsx'


def write_row(worksheet, col_number, v, detect=False):
    step = 0
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=col_number, max_col=col_number):
        step += 1
        if detect:
            if worksheet.cell(step + 1, col_number).value is None:
                worksheet.cell(step + 1, col_number).value = v
    return col_number


def open_file(file_path, col_number, v, detect):
    # 打开现有的 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 创建一个新的 Excel 文件
    worksheet = workbook.active
    write_row(worksheet, col_number, v, detect)
    workbook.save(path)
    workbook.close()


def order(file_path, v):
    # 打开工作簿
    wb = openpyxl.load_workbook(file_path)
    # 选择第一个工作表
    ws = wb.active
    # 创建一个空列表来存储数据
    data_list = v
    # 扫描第一列找到空的单元格并添加数据
    step = 0
    while True:
        step +=1
        if ws.cell(step, 1).value is None:
                n = 0
                for i in data_list:
                    n += 1
                    ws.cell(step, n).value = i
                break

    # 关闭工作簿
    wb.save(file_path)
    wb.close()
    # 输出列表数据
    print(data_list)
# buy_number: 4
# desc: "灵力价"
# goods_price: 50
# goods_title: "永诺YN85mm F1.8S（金属外壳版）索尼E口全画幅定焦自动对焦镜头"
# notes: ""
# phone_number: "17777253073"
# tag: "小A"
